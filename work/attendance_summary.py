import psycopg2
import pandas as pd
import re
from datetime import datetime
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from holidays import HOLIDAYS, MONTH, get_working_days, YEAR
import os
from config import DB_CONFIG
import sys
import calendar

days_in_month = calendar.monthrange(YEAR, int(MONTH))[1]

def flush_print(*args, **kwargs):
    print(*args, **kwargs)
    sys.stdout.flush()

OUTPUT_DIR = "output"
os.makedirs(OUTPUT_DIR, exist_ok=True)
timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
output_file = os.path.join(OUTPUT_DIR, f"考勤明细及统计_{timestamp}.xlsx")

def get_db_connection():
    return psycopg2.connect(**DB_CONFIG)

def count_attendance_status(row):
    counts = {
        "正常次数": 0,
        "迟到次数": 0,
        "早退次数": 0,
        "缺卡次数": 0,
        "旷工次数": 0,
        "出差次数": 0,
        "请假次数": 0,
        "钉钉加班时长(h)": 0.0,
        "飞书加班时长(h)": 0.0,
        "总加班时长(h)": 0.0,
        "迟到具体日期": [],
        "早退具体日期": [],
        "缺卡具体日期": [],
        "旷工具体日期": [],
        "出差具体日期": [],
        "请假具体日期": [],
        "同行出差次数": 0,
        "同行出差具体日期": [],
        "发起出差次数": 0,
        "发起出差具体日期": [],
        "事假次数": 0,
        "病假次数": 0,
        "婚假次数": 0,
        "产假次数": 0,
        "陪产假次数": 0,
        "丧假次数": 0,
        "调休次数": 0,
    }

    leave_types = ["事假", "病假", "婚假", "产假", "陪产假", "丧假", "调休"]

    for day in range(1, days_in_month + 1):
        day_col = f"第{day}天"
        if day_col not in row:
            continue
        status = str(row[day_col])
        if not status or status == 'nan':
            continue
        day_str = f"{day:02d}"
        if "正常" in status:
            counts["正常次数"] += 1
        if "迟到" in status:
            counts["迟到次数"] += 1
            counts["迟到具体日期"].append(str(day))
        if "早退" in status:
            counts["早退次数"] += 1
            counts["早退具体日期"].append(str(day))
        if "缺卡" in status:
            # 提取缺卡次数
            match = re.search(r"缺卡\((\d+)次\)", status)
            if match:
                missing_count = int(match.group(1))
                counts["缺卡次数"] += missing_count
            else:
                # 兼容旧格式 "缺卡(1天)"
                counts["缺卡次数"] += 1
            counts["缺卡具体日期"].append(str(day))
        if "旷工" in status:
            counts["旷工次数"] += 1
            counts["旷工具体日期"].append(str(day))
        if "出差" in status:
            counts["出差次数"] += 1
            counts["出差具体日期"].append(str(day))
            # 检查是否有同行人信息（发起人）
            if "[同行人:" in status:
                counts["发起出差次数"] += 1
                counts["发起出差具体日期"].append(str(day))
            # 检查是否是同行出差（被发起人）
            elif "[发起人:" in status:
                counts["同行出差次数"] += 1
                counts["同行出差具体日期"].append(str(day))
        if "请假" in status and day_str not in HOLIDAYS:
            counts["请假次数"] += 1
            counts["请假具体日期"].append(str(day))
            for leave_type in leave_types:
                if leave_type in status:
                    counts[f"{leave_type}次数"] += 1
        overtime_matches = re.findall(r'(钉钉加班|飞书加班)\((\d+\.?\d*)h\)', status)
        for source, hours in overtime_matches:
            try:
                hours_float = float(hours)
                if source == "钉钉加班":
                    counts["钉钉加班时长(h)"] += hours_float
                    counts["总加班时长(h)"] += hours_float
                elif source == "飞书加班":
                    counts["飞书加班时长(h)"] += hours_float
                    counts["总加班时长(h)"] += hours_float
            except ValueError:
                flush_print(f"警告: 无法转换加班时长 '{hours}' 为数字")
                continue

    for key in ["迟到具体日期", "早退具体日期", "缺卡具体日期", "旷工具体日期", "出差具体日期", "同行出差具体日期", "发起出差具体日期", "请假具体日期"]:
        counts[key] = ",".join(counts[key]) if counts[key] else ""
    return pd.Series(counts)

def format_attendance_status(status, day):
    if not status or str(status) == 'nan':
        return ''
    status = str(status)
    formatted = []
    day_str = f"{day:02d}"
    is_holiday = day_str in HOLIDAYS
    if is_holiday:
        formatted.append("🏠")
    if "正常" in status:
        formatted.append("✅")
    if "迟到" in status:
        formatted.append("⏰")
    if "早退" in status:
        formatted.append("⚡")
    if "缺卡" in status:
        formatted.append("❌")
    if "旷工" in status:
        formatted.append("⛔")
    if "出差" in status:
        if "[同行人:" in status:
            formatted.append("🚗👥")  # 发起出差，带同行人
        elif "[发起人:" in status:
            formatted.append("👥🚗")  # 同行出差，被发起
        else:
            formatted.append("🚗")    # 普通出差
    if "请假" in status:
        formatted.append("📝")
    if formatted:
        if is_holiday:
            return f"{''.join(formatted)} 休息日\n{status}"
        else:
            return f"{''.join(formatted)} {status}"
    return status

def analyze_attendance():
    conn = get_db_connection()
    try:
        query = "SELECT * FROM attendance_result"
        df = pd.read_sql_query(query, conn)
        df = df.sort_values(by='部门', ascending=True)

        statistics = df.apply(count_attendance_status, axis=1)
        df_stats = pd.concat([df[['姓名', '考勤组', '部门', '工号', '职位', 'UserId']], statistics], axis=1)
        working_days = get_working_days()
        df_stats['应出勤天数'] = int(working_days)
        df_stats['缺卡次数'] = df_stats['缺卡次数'].astype(int)
        # 修改实际出勤天数的计算逻辑，改为正常出勤次数加上出差次数
        df_stats['实际出勤天数'] = df_stats['正常次数'] + df_stats['出差次数']

        base_info_cols = ['姓名', '考勤组', '部门', '工号', '职位', 'UserId']
        statistic_cols = [
            "事假次数", "病假次数", "婚假次数", "产假次数", "陪产假次数", "丧假次数", "调休次数",
            "正常次数", "迟到次数", "早退次数", "缺卡次数", "旷工次数", "出差次数", "发起出差次数", "同行出差次数", "请假次数",
            "钉钉加班时长(h)", "飞书加班时长(h)", "总加班时长(h)",
            "应出勤天数", "实际出勤天数",
            "迟到具体日期", "早退具体日期", "缺卡具体日期", "旷工具体日期", "出差具体日期", "发起出差具体日期", "同行出差具体日期", "请假具体日期"
        ]
        other_cols = [col for col in df_stats.columns if col not in base_info_cols + statistic_cols]
        df_stats = df_stats[base_info_cols + statistic_cols + other_cols]

        for day in range(1, days_in_month + 1):
            day_col = f"第{day}天"
            if day_col in df.columns:
                df[day_col] = df.apply(lambda row: format_attendance_status(row[day_col], day), axis=1)

        long_records = []
        for _, row in df.iterrows():
            base_info = {col: row[col] for col in base_info_cols}
            for day in range(1, days_in_month + 1):
                col_name = f'第{day}天'
                if col_name not in df.columns:
                    continue
                status = row[col_name]
                if pd.isna(status):
                    status = ''
                record = base_info.copy()
                record['日期'] = f"{MONTH}-{day:02d}"
                record['状态'] = status
                long_records.append(record)
        df_long_detail = pd.DataFrame(long_records)

        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            df_stats.to_excel(writer, index=False, sheet_name='考勤统计')
            df_long_detail.to_excel(writer, index=False, sheet_name='考勤明细（纵向）')
            df.to_excel(writer, index=False, sheet_name='考勤明细')

            workbook = writer.book

            # 格式化“考勤明细”和“考勤明细（纵向）”
            for sheet_name in ['考勤明细', '考勤明细（纵向）']:
                worksheet = writer.sheets[sheet_name]
                header_font = Font(name='微软雅黑', bold=True, size=11, color='000000')
                header_fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
                border = Border(
                    left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin')
                )
                header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                for cell in worksheet[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.border = border
                    cell.alignment = header_alignment
                worksheet.freeze_panes = 'B2'
                for col_cells in worksheet.columns:
                    length = max(len(str(cell.value)) if cell.value else 0 for cell in col_cells)
                    col_letter = get_column_letter(col_cells[0].column)
                    worksheet.column_dimensions[col_letter].width = min(length + 5, 40)
                data_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                for row in worksheet.iter_rows(min_row=2):
                    for cell in row:
                        cell.alignment = data_alignment
                        cell.border = border

            # 格式化“考勤统计”
            worksheet = writer.sheets['考勤统计']

            header_font = Font(name='微软雅黑', bold=True, size=11, color='000000')
            header_fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
            border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            # 表头格式
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
                cell.alignment = header_alignment

            worksheet.freeze_panes = 'B2'

            # 前6列自动宽度计算
            for col_idx in range(1, 7):
                col_cells = list(worksheet.iter_cols(min_col=col_idx, max_col=col_idx, min_row=1, max_row=worksheet.max_row))[0]
                length = max(len(str(cell.value)) if cell.value else 0 for cell in col_cells)
                col_letter = get_column_letter(col_idx)
                worksheet.column_dimensions[col_letter].width = min(length + 5, 40)

            # 第7列及以后列宽固定10
            fixed_width_rest = 10
            max_col = worksheet.max_column
            for col_idx in range(7, max_col + 1):
                col_letter = get_column_letter(col_idx)
                worksheet.column_dimensions[col_letter].width = fixed_width_rest

            # 所有行高统一20（包含第一行）
            fixed_row_height = 20
            for row_idx in range(1, worksheet.max_row + 1):
                worksheet.row_dimensions[row_idx].height = fixed_row_height

            # 数据单元格格式（第一行已处理）
            data_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            for row in worksheet.iter_rows(min_row=2):
                for cell in row:
                    cell.alignment = data_alignment
                    cell.border = border

        flush_print(f"✅ 考勤数据已导出到: {output_file}")
        flush_print("\n📊 考勤统计概览:")
        flush_print(f"总人数: {len(df)}")
        flush_print(f"应出勤天数: {working_days}")
        flush_print(f"平均实际出勤天数: {df_stats['实际出勤天数'].mean():.1f}")
        return df

    except Exception as e:
        flush_print(f"❌ 错误: {e}")
        return None
    finally:
        conn.close()

def main():
    df = analyze_attendance()
    if df is not None:
        flush_print("\n✅ 考勤处理完成")

if __name__ == "__main__":
    main()
